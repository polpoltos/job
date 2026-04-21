from openpyxl import load_workbook

# имя файла и лист
filename = "адреса.xlsx"
sheet_name = "адреса"  # поменяй, если лист называется иначе

wb = load_workbook(filename)
ws = wb[sheet_name]

for row in range(1, 3001):  # A1-A3000
    cell = ws[f"A{row}"]
    if isinstance(cell.value, str):
        cell.value = cell.value.replace(" ", "").replace("\t", "")

wb.save(filename)
print("Готово: пробелы в A1-A3000 удалены.")
